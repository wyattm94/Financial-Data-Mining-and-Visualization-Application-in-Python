# -*- coding: utf-8 -*-
"""
__________________________________________________________________________________________________

    > Script ID     : helper.py
    > Author(s)     : Wyatt Marciniak [wyattm94@gmail.com]
    > Maintainer(s) : Wyatt Marciniak
    > License(s)    : Copyright (C) Wyatt Marciniak - All Rights Reserved [See LICENSE File]
    > Description   : Helper Functions (assoc. web scraping APIs)
__________________________________________________________________________________________________
<<< Source Code Begins >>>
"""

import os
import sys
import time
import numpy as np
import pandas as pd

import datetime as dt
from datetime import datetime as dtt

from bs4 import BeautifulSoup
import requests as req
import xlsxwriter
import math
import gc
import random as rand
import openpyxl as opx
import threading
from threading import Thread as mt
from multiprocessing.pool import ThreadPool as mtpool
from multiprocessing import Process, Queue
from threading import active_count as nlive
import certifi


# region > Optimal (ID to Save)

def crt(t0,sigfig=3,scale=1): return np.round((time.time()-t0)/scale,sigfig)

def isok(x):
    """ Returns True if object is not None type """
    return (x is not None)

def conv_s2n(x,to_ignore=None):
    """ Handles string to number conversions (Biased towards Financial Data"""
    if isinstance(x,(list,)): #and len(x) > 1:
        nds = []
        for e in x:
            nds.append(conv_s2n(e,to_ignore=to_ignore))
        return nds
    else:
        if to_ignore is not None:
            for i in to_ignore:
                if x.find(i) > -1: return x

        if x is None: return x
        if x in ['na','NA','n/a','N/A','N/a','n/A','NE']: return None
        if x in ['-','--']: return 0
        if x in ['unch','Unch','UNCH']: return float(0)

        if type(x) is float or type(x) is int or \
                str(type(x)) == "<class 'numpy.float64'>" or \
                str(type(x)) == "<class 'numpy.int32'>" or \
                str(type(x)) == "<class 'numpy.int64'>":
            return float(x)

        x = x.replace(',','')
        x = x.replace('+','')
        x = x.replace('$','')

        if x.find('T') > -1: return float(x.replace('T',''))*1000
        if x.find('M') > -1: return float(x.replace('M', ''))*1000000
        if x.find('B') > -1: return float(x.replace('B', ''))*1000000000

        tpos = x.find('%')
        if tpos != -1: return float(x[0:tpos]) / 100
        elif float_ok(x): return float(x)
        else: return x

# [USE] Fastest Method for reading xlsx files into Python (Uses Multi-processing on a Thread Pool)
def wb2df(wb):
    cv = wb.values; cn = next(cv)[0:]
    return pd.DataFrame(cv, columns=cn)

def xb2py(p,mtmax=10):
    gc.collect(); t0 = time.time(); out = dict()

    # Open connection to xlsx file
    wb = opx.load_workbook(p); ws = list(wb.sheetnames)

    # Pass args to pool / terminate and clean
    pool = mtpool(mtmax); out = pool.map(wb2df,[wb[s] for s in ws])
    pool.terminate(); wb.close(); out = mkdict(ws,out); gc.collect()

    # Handle Returns
    if len(out.keys()) == 1: return out[list(out.keys())[0]]
    else: return out



# endregion

# region > FINAL APP
def mklist(x): return [a for a in x]

def istype(x, *args):
    # return args
    test = mklist(list(args));
    types, ntypes = [], []
    types = [a for a in test if str(type(a)) == "<class 'type'>"]
    ntypes = [a for a in test if str(type(a)) != "<class 'type'>"]
    if len(ntypes) > 0:
        sysout('\n[WARN] Non-types entered = ', len(ntypes), ':\n')
        print(ntypes, ' ')
    return bool(type(x) in types)

def make_slice(x,frc=1):
    if isinstance(x, (int, float)):
        return slice(int(x), int(x) + 1,frc)
    elif isinstance(x, tuple):
        return slice(x[0], x[1],frc)
    elif isinstance(x, range):
        # print(slice(list(x)[0], list(x)[(-1)] + 1))
        return slice(list(x)[0], list(x)[(-1)] + 1,frc)
    else:
        return [make_slice(e,frc) for e in x]

def slicer(s, d, bycol=False):
    slices = [make_slice(x) for x in s]
    print(slices)
    if istype(d, list):
        return unlist([d[x] for x in list(slices)])
    elif bycol:
        k = unlist(slicer(s, list(d.columns)))
        v = []
        for x in slices:
            temp = d.iloc[:,x]
            if len(temp.columns) > 1:
                for cn in temp.columns:
                    v.append(list(temp[cn]))
            else: v.append(list(temp[list(temp.columns)[0]]))
        return kv2df(k=k,v=v)
    else:
        v = d[slices[0]]
        print(v)
        if len(slices) == 1:
            return v
        else:
            for x in slices[1:]: v = v.append(d[x])
            return v

def unlist(x):
    out = []
    for e in x:
        if istype(e, list, tuple):
            out += [a for a in e]
        elif istype(e, int, float, str):
            out += [e]
    return out

# endregion


# def utime(nice=True):
#     orig = int(dt_tag().split('_')[1])
#     if nice:
#         return {'orig': orig, 'nice': nicetime(orig)}
#     else:
#         return orig
#
# def nicetime(x):
#     x = ifelse(len(str(x))<6,str(str(0)+str(x)),str(x))
#     return mkstr(x[0:2],x[2:4],x[4:6],sep=':')
#
# def nicedate(x):
#     return mkstr(['20',x[:2]],'',x[2:4],x[4:6],sep='-',sep_list='')

# def ftime(x):
#     t0 = utime()
#     op = eval(x)
#     t1 = utime()
#     t0_a = nicetime(t0)
#     t1_a = nicetime(t1)
#     return {'op':op,'to_n':t0,'t1_n':t1,'diff':t1-t0}

def id_row(df,col,rval,num=False):
    '''  '''
    if num:
        ce = [i for i,e in enumerate(df[col]) if e == rval]
        ce = ifelse(len(ce)==0,[-1],ce)

        cg = [i for i,e in enumerate(df[col]) if e > rval]
        if len(cg) == 0 and len(ce) == 1: cg = ce
        else: cg = ifelse(len(cg) == 0, [-1], cg)

        cl = [i for i,e in enumerate(df[col]) if e < rval]
        if len(cl) == 0 and len(ce) == 1: cl = ce
        else: cl = ifelse(len(cl) == 0, [-1], cl)
        # cl = ifelse(len(cl) == 0, [-1], cl)

        return {'e':ce,'g':cg,'l':cl}

    else:
        chk = [i for i,e in enumerate(df[col]) if e.lower() == rval.lower()]
        if len(chk) == 0: return [-1]
        else: return chk

# def rbind(to_rbind,df=None):
#     if df is None: df = pd.DataFrame()
#     for r in to_rbind: df = df.append([r],ignore_index=False)
#     return df
#
# def cbind(df,add):
#     t = list(df.columns) + [k for k in list(add.keys())]
#     v = [df[c] for c in t[:len(df.columns)]] + [add[k] for k in list(add.keys())]
#     return(pd.DataFrame(dict(zip(t,v))))

def aggregate(df):
    out = []
    for cn in list(df.columns): out += list(df[cn])
    return out



def float_ok(x):
    try:
        float(x)
        return True
    except ValueError:
        return False

def dt_tag():
    raw  = str(str(dtt.today()).split('.')[:-1]).split(' ')
    date = raw[0].strip("[\\'").replace('-','')[2:]
    time = raw[1].strip("\\']").replace(':','')
    return str(date+'_'+time)

def mkfpath(d=None,f=None,e=None,tag=False):
    if f is None: return None
    elif e is None: e = '.xlsx'
    elif d is None: return str(f+e)
    else: return str(d+'/'+f+e)

def ifelse(c,t,f):
    if bool(c) is True : return t
    else: return f

def isin(is_,in_,only_vals=False):
    # if np.unique(in_) != in_: in_ = np.unique(in_)
    # if np.unique(is_) != is_: is_ = np.unique(is_)
    tf_iso   = [[a == b for a in is_] for b in in_]
    tf_count = [str(x).count('True') for x in tf_iso]#list(tf_iso.items())]
    tf_isin  = [float(x) > 0 for x in tf_count]
    tf_id    = [i for i in range(0,len(tf_isin)) if tf_isin[i] == True]
    # tf_val   = [new_name[i] for i in range(0,len(new_name)) if tf_isin[i] == True]
    tf_val = [in_[i] for i in range(0, len(tf_isin)) if tf_isin[i] == True]

    if only_vals: return tf_val #return {'v':tf_val,'id':tf_id}
    return {'iso':tf_iso,'count':tf_count,'isin':tf_isin,'v':tf_val,'id':tf_id,
            'base':in_,'chk':is_}

def tracer(t=False,*args,sep=''):
    if t: print(mkstr(args,sep))



def concat_list(list,sep=' '):
    s = ''
    for i,e in enumerate(list):
        if i == 0: s += str(e)
        else: s += str(sep) + str(e)
    return s

def mkstr(*args,sep='',sep_list='',sep_tuple=' '):
    s,use_sep = '',''
    for i in range(len(args)):
        a = args[i]
        use_sep = ifelse(bool(isinstance(a,list)),sep_list,
                         ifelse(bool(isinstance(a,tuple)),sep_tuple,sep))
        if not isinstance(a,(list,tuple)): s += str(a) + ifelse(bool(i < len(args)),str(use_sep),'')
        else:
            for e in a: s += str(e) + ifelse(bool(i < len(args)),str(use_sep),'')
    if s[len(s)-len(use_sep):]==use_sep: return s[:len(s)-len(use_sep)]
    else: return s
    # s = s[:-1]

def sysout(*args,sep=('','',' '),log=False):
    out = mkstr(*args,sep=sep[0],sep_list=sep[1],sep_tuple=sep[2])
    if log: p2log(out)
    sys.stdout.write(out); sys.stdout.flush()

def repeat(r,t,sep=None):
    if isok(sep):
        return mkstr([r for i in range(t)],sep_list=sep)
    else: return [r for i in range(t)]

def p2log(*args,fp='.logfile',local=False):
    tag = mkstr('\n',repeat('-',25,''),'\n[User Log Updated: ',str(dtt.today()),']\n----\n')
    pconn = open(fp,'a+')
    print(tag, file=pconn)
    for a in args:
        print(a,file=pconn)
    print('\n-----[END]-----\n', file=pconn)
    pconn.close()
    if local: print(*args)

# p2log([1,2,3,4,5],{'a':1,'b':2},fp='test.logfile')


def df2dict(df):
    k = list(df.columns)
    v = [df[e] for e in k]
    return mkdict(k,v)

# def apply_ranks(d,col_tag=0,col_rank=None,asc=False):
#     coln = list(d.columns)
#     tags = d.iloc[:,col_tag]
#     d_upd = d.copy()
#     if isok(col_rank):
#         c_rank = [coln[col_tag]]
#         slices = make_slice(col_rank)
#         for s in slices:
#             c_rank += coln[s]
#         d_upd.drop([c for c in coln if c not in c_rank],axis=1,inplace=True)
#
#     # Iterate over the columns, get ranks and add (matched to tag col)
#     ranks = dict()
#     for c in d_upd.columns[1:]:
#         # Create is DF, sort by target col, add 1 -> n ranks, reorder index, save ranks
#         temp = pd.DataFrame(dict(zip(['tags',c],([tags]+[d_upd[c]]))))
#         temp.sort_values(by=c,axis=0,ascending=asc,inplace=True)
#         temp = cbind(temp,{'ranks': [i for i in range(1, len(temp.index) + 1)]})
#         temp.sort_index(axis=0,ascending=True,inplace=True)
#         ranks[mkstr('rank_',c)] = temp['ranks']
#     return cbind(d,ranks)
#
#         # print(coln[np.where(coln != )])
#         # to_drop = coln[col_rank]
#
#
#     # d.sort_values()
#     # print(coln)
#     # print(d.columns)

def l_(x):
    if isinstance(x,(list,)):
        return [l_(e) for e in x]
    else: return x.lower()

def u_(x):
    if isinstance(x,(list,)):
        return [u_(e) for e in x]
    else: return x.upper()

# def wbget(u,f1_n=None,f1_p=None,f1_t=None):
#     raw = BeautifulSoup(req.get(u).content, features='html.parser')
#     to_eval = str('raw')
#     if f1_n is not None:
#         to_eval = str(to_eval+'.'+f1_n+'(')
#         if f1_p and f1_t is not None :
#             to_eval = str(to_eval+f1_p+'=')
#         if f1_t is not None:
#             to_eval = str(to_eval+"'"+f1_t+"'")
#         to_eval = str(to_eval+')')
#         return eval(to_eval)
#     else:
#         return raw

def ws_raw2df(x,numcols=[],xc_writer=None,fp=None,sn=None,index=False,sort=None,local=False):
    # Extract inputs, if colvals is not a list of lists, make it one
    newdict = {}; colnames = x['head']; colvalues = x['body']
    if not isinstance(colvalues,(list,)): colvalues = [list(cv) for cv in colvalues.values()]

    # Adjust value size to match colnames size (can be changed)
    if len(colnames) < len(colvalues):
        colnames.append(str('col_'+str(x)) for x in range(len(colnames),len(colvalues)))
    for i in range(0,len(colnames)):
        if i > 0 and len(colvalues[i]) < len(colvalues[(i-1)]): continue  #Omit cols w/missing data
        if colnames[i] is None: colnames[i] = str('col_'+str(i+1))

        # Add and filter (str -> float) if specified in numcols
        if i in numcols: newdict[colnames[i]] = conv_s2n(colvalues[i])
        else: newdict[colnames[i]] = colvalues[i]

    df = pd.DataFrame(newdict)
    if sort is not None: df.sort_values(by=[sort],ascending=False) # Order dates by unix stamps

    # Outputs
    if local: return df  # If local, only local
    elif xc_writer is not None: df.to_excel(xc_writer, sheet_name=sn, index=index)
    elif fp is not None: df.to_excel(fp, sheet_name=sn, index=index)
    else: sysout('\n> [!!!] Unknown Error\n')

def wb2dict(p_wb):
    return {x:pd.read_excel(p_wb,x) for x in list(pd.ExcelFile(p_wb).sheet_names)}

def dict2wb(p_wb,d,try_k2n=True):
    temp_writer = pd.ExcelWriter(p_wb)
    for i,s in enumerate(list(d.keys())):
        temp = ifelse(len(s)>30 or not try_k2n,
                      str('dataset_'+str(i+1)),s)
        try:
            d[s].to_excel(temp_writer, sheet_name=temp, index=False)
        except Exception: continue
    temp_writer.save(); temp_writer.close()

def kv2df(k,v):
    v = [ifelse(bool(isinstance(x,list)),x,[x]) for x in v]
    return pd.DataFrame(mkdict(k,v))

def mkdict(k,v):
    return dict(zip(k,v))
    # out = dict()
    # for i in range(len(k)):
    #     out[k[i]] = v[i]
    # return out

# Add *args that are used (in order) to parse a dictionary, list, anything...
# def pcf(x,*args):
#     tx = x
#     for a in list(args): tx = tx[a]
#     return tx

# Determine T/F by checking wide (use with pcf)
def get_tf(b):
    return b in ('TRUE','True','true','T','t',
                 'YES','Yes','yes','Y','y',
                 'OK','Ok','ok')

# Set up Config settings + Local copy of config file
# def load_config(p='.app_config',ret_local=False):
#     # Create Global 'local' config file + Read project config file
#     if ret_local: cflocal = dict()
#     else: globals()['cflocal'] = dict()
#
#     # Open/load config file, set collectors ([]) + read lines into a list for iteration
#     temp_cf = open('.app_config', 'r')
#     keys = []; vals = []; key0 = None
#     t2 = temp_cf.readlines()
#
#     # Iterate and populate the dict hierarchally
#     for e in t2:
#         # Split into word list --> Check for cases
#         # ('_' = a partition), append collected data (under new head key)
#         elems = e.split(', ')
#         if len(elems) == 1:
#             cflocal[key0] = dict(zip(keys, vals))
#             keys = []; vals = []; key0 = None
#         else:
#             if key0 is None: key0 = elems[0].replace('\n','')
#             keys.append(str(elems[1]).replace('\n',''))
#             vals.append(str(elems[2]).replace('\n',''))
#
#     # Check if backend dir/files are setup
#     # if get_tf(pcf(cflocal,'flag','backend_ok')): return
#
#     # Create Directories
#     p_warehouse = pcf(cflocal,'path','warehouse')
#     for k in [p_warehouse]+[str(p_warehouse+x) for x in ['equity/','etf/','fx/']]:
#         if not os.path.exists(k): os.mkdir(k)
#
#     # Create Log File (Excel wb - sheets)
#     fn_logfile = pcf(cflocal, 'path', 'logfile')
#     writer = pd.ExcelWriter(fn_logfile) # Connection -> write only
#     coln = ['ticker','updated_u','updated_c','user','fpath'] # Colnames
#     # Check logfile -> check exist and for proper sheetnames
#     if not os.path.exists(fn_logfile):
#         logdict = dict(zip(['equity', 'etf', 'fx'],
#                            [pd.DataFrame(columns=coln) for i in range(0,3)]))
#         dict2wb(writer, logdict)
#         sysout('\n> Log File Created\n-----\n\n')
#     else:
#         temp = wb2dict(writer)
#         temp_sn = list(temp)
#         check = all(x in temp_sn for x in ['equity','etf','fx'])
#         if check: sysout('\n> Log File Found\n-----\n\n')
#         else:
#             logdict = dict(zip(['equity', 'etf', 'fx'],
#                                [pd.DataFrame(columns=coln) for i in range(0, 3)]))
#             dict2wb(writer, logdict)
#             sysout('\n> Log File Created\n-----\n\n')
#
#     # Create Cache (if flag is T)
#     if get_tf(pcf(cflocal,'opts','use_cache')):
#         if 'cache' not in globals():
#             global cache; cache = {}
#             sysout('\n> Cache Created\n-----\n\n')
#         else:
#             sysout('\n> Cache Found\n-----\n\n')
#
#     if ret_local: return cflocal, {}

# Create a list of dict entries (label:value) for Dash Applications (Wrapper)
def create_options(l,v):
    return list({'label':l[i],'value':v[i]} for i in range(0,len(v)))

# Checking Caching Times - File sizes/sheets/pull_times (see root)
def cache_timing(name='test1',fn='cache_time_test.xlsx',rdata=False):
    data = {}
    dcols = ['ac','ticker','file','size','sheets','time']
    ac_l,t_l,file_l,size_l,sheets_l,time_l = ([] for i in range(0,6))
    # stats = pd.DataFrame(columns=dcols)
    for ac in os.listdir('warehouse'):
        sysout('> Asset Class: ',ac,'\n')
        path = str('warehouse/'+ac)
        for t in os.listdir(path):
            # Path + isolated ticker
            path2 = str(path+'/'+t)
            t_iso = str(t.split('.')[0])
            sysout('  + Ticker: ',t_iso,)
            # Run + timing + extract/store stats
            t0 = time.time(); temp = wb2dict(path2); t1 = time.time()
            data[t_iso] = temp

            ac_l.append(ac)
            t_l.append(t_iso)
            file_l.append(t)
            size_l.append(np.round(float(os.path.getsize(path2))/1000,2))
            sheets_l.append(len(list(temp)))
            time_l.append(float(np.round(t1-t0,3)))

            sysout(' - Time: ',float(np.round(t1-t0,3)),' sec.\n')
    dvals = [t_l,ac_l,file_l,size_l,sheets_l,time_l]
    pd.DataFrame(dict(zip(dcols,dvals))).to_excel(fn,name,index=False)
    if rdata: return data

# Function termination helper (sysout() + return None)
def end(s='[Function end() called] Current function terminated'):
    sysout(mkstr('\n',s)); return None

# Quick rouding wrapper for structures (including single numbers)
def roundwrap(x,sigfig=4):
    if isinstance(x,list): return [round(a,sigfig) for a in x]
    elif isinstance(x,dict): return 1
    elif isinstance(x,pd.DataFrame): return 1
    else: return round(x,sigfig)

# Find rows in DF (return loc or subset DF)
def idrows(df,det,cond,val,as_df=False):
    try:
        if not isinstance(df,pd.DataFrame): return end('[ERROR] df not pd.DataFrame')
        if not isinstance(det, (str,int,float)): return end('[ERROR] by not (str,int,float)')
        if not isinstance(det, (str,int,float)): return end('[ERROR] by not (str,int,float)')
    except IOError: return end('[ERROR] Missing/Bad param inputs')
    else:
        d = None
        if isinstance(det,(int,float)):  id = [list(df.columns)[i] for i in det]
        else: id = det
        if isinstance(eval('val'),str):  d = [str(round(x,3)) for x in df[id]]
        else: d = [float(x) for x in df[id]]
        # print(mkstr(round(d[0],3),cond, val))
        rows = [i for i,x in enumerate(d) if eval(mkstr(round(x,3),cond, val)) is True]
        if as_df: return df.iloc[rows]
        return rows

# [X] Excel Workbook to DF --> Much faster than pd.read.excel() (wb2df)
def  xb2df(conn):
    dset = dict()
    dconn = opx.load_workbook(conn)
    for sn in dconn.sheetnames:
        raw_data = dconn[sn].values
        colnames = next(raw_data)[0:]
        dset[sn] = pd.DataFrame(raw_data, columns=colnames)
    dconn.close()
    return dset



# Dictionary Quick-Parser
def usekeys(d,k,keep_bad=False):
    loc = None
    dk = list(d.keys())
    if not isinstance(k,list):
        try:
            loc = np.where(np.array(dk)==k)[0]
            if isok(loc): return d[dk[int(loc)]]
        except TypeError:
            sysout('\n[!] Bad Key: ',k); return None
    else:
        out = dict()
        for e in k:
            if e in out.keys(): continue
            temp = usekeys(d, e)
            if keep_bad or isok(temp) : out[e] = temp
        return out


# time.time()

def showdt(as_str=False):
    s = str(dtt.utcfromtimestamp(time.time()))
    if as_str: return s.split('.')[0] + mkstr('.', s.split('.')[1][:3])
    else: return s.split(' ')[0], s.split(' ')[1].split('.')[0] + mkstr('.', s.split('.')[1][:3])
# showdt()